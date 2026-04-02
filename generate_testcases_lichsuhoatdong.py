import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Cases Lịch sử hoạt động"

# Styles
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='top')
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, size=11, color="FFFFFF")

gui_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
gui_font = Font(bold=True, size=11, color="1F4E79")

loc_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
loc_font = Font(bold=True, size=11, color="375623")

pvtd_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
pvtd_font = Font(bold=True, size=11, color="2E7D32")

gtb_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
gtb_font = Font(bold=True, size=11, color="E65100")

pvtd_dl_fill = PatternFill(start_color="E0F7FA", end_color="E0F7FA", fill_type="solid")
pvtd_dl_font = Font(bold=True, size=11, color="006064")

gtb_dl_fill = PatternFill(start_color="FBE9E7", end_color="FBE9E7", fill_type="solid")
gtb_dl_font = Font(bold=True, size=11, color="BF360C")

# Headers
headers = [
    "ID\n(Mã chức năng)", "Test Title", "Test Case Description\n(Mô tả trường hợp)",
    "Test Case Procedure\n(Các bước thực hiện)", "Test Data\n(Dữ liệu truyền vào)",
    "Expected Output\n(Kết quả mong muốn)", "Actual Result\n(Kết quả thực tế)",
    "Video minh chứng", "Inter-test case\nDependence", "Result 1",
    "Test date 1\n(Ngày thực hiện)", "Result 2", "Test date 2\n(Ngày thực hiện)",
    "Video minh chứng", "Người thực hiện"
]

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    cell.border = thin_border

col_widths = [16, 45, 50, 60, 45, 50, 50, 30, 20, 10, 15, 10, 15, 30, 15]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

def write_section(ws, row, title, fill, font):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = font
    cell.fill = fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    for c in range(1, 16):
        ws.cell(row=row, column=c).border = thin_border
        ws.cell(row=row, column=c).fill = fill
    return row + 1

def write_tests(ws, row, tests):
    for test in tests:
        for col, val in enumerate(test, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = wrap_align
            cell.border = thin_border
        row += 1
    return row

row = 2
steps_base = 'Bước 1: Đăng nhập thành công vào hệ thống SevenStrike với vai trò "Admin"\nBước 2: Truy cập vào màn "Lịch sử hoạt động" (trong module Lịch làm việc)\n'

# ============ GUI ============
row = write_section(ws, row, "KIỂM TRA GUI LỊCH SỬ HOẠT ĐỘNG", gui_fill, gui_font)

gui_tests = [
    ["LSHD_GUI_01", "Kiểm tra hiển thị màn hình Lịch sử hoạt động",
     "Kiểm tra hiển thị tổng thể giao diện màn hình Lịch sử hoạt động",
     steps_base + "Bước 3: Kiểm tra hiển thị giao diện màn hình Lịch sử hoạt động", "N/a",
     "Hiển thị đầy đủ các thành phần: tiêu đề, khu vực bộ lọc, bảng danh sách lịch sử",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_GUI_02", "Kiểm tra hiển thị tiêu đề",
     'Kiểm tra hiển thị tiêu đề "Lịch sử hoạt động" đúng chính tả và vị trí',
     steps_base + "Bước 3: Kiểm tra hiển thị tiêu đề màn hình", "N/a",
     'Hiển thị tiêu đề "Lịch sử hoạt động" đúng chính tả, đúng vị trí',
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_GUI_03", "Kiểm tra hiển thị khu vực bộ lọc",
     "Kiểm tra hiển thị đầy đủ các trường trong khu vực bộ lọc tìm kiếm",
     steps_base + "Bước 3: Kiểm tra hiển thị khu vực bộ lọc", "N/a",
     "Hiển thị đủ các trường: Ô tìm kiếm, Từ ngày, Đến ngày",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_GUI_04", "Kiểm tra hiển thị placeholder ô tìm kiếm",
     "Kiểm tra hiển thị placeholder trong ô tìm kiếm",
     steps_base + "Bước 3: Kiểm tra hiển thị placeholder ô tìm kiếm", "N/a",
     "Hiển thị placeholder đúng nội dung, đúng chính tả, đúng vị trí",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_GUI_05", "Kiểm tra hiển thị tiêu đề các cột trong bảng",
     "Kiểm tra hiển thị tiêu đề các cột trong bảng Lịch sử hoạt động",
     steps_base + "Bước 3: Kiểm tra hiển thị tiêu đề các cột trong bảng", "N/a",
     "Hiển thị các cột: STT, Tài khoản, Ca, Mở, Đóng, Tiền mặt, Chênh lệch, Trạng thái",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_GUI_06", "Kiểm tra hiển thị badge ca làm",
     "Kiểm tra hiển thị badge ca làm đúng màu sắc theo loại ca (Sáng/Chiều/ADMIN)",
     steps_base + "Bước 3: Kiểm tra hiển thị badge ca làm trong cột Ca", "N/a",
     "Hiển thị badge ca làm đúng màu: ca Sáng, ca Chiều có màu riêng, ADMIN có badge riêng biệt",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_GUI_07", "Kiểm tra hiển thị trạng thái dạng badge",
     'Kiểm tra hiển thị trạng thái dạng badge với các giá trị "Đang làm", "Đã đóng", "Hoàn tất"',
     steps_base + "Bước 3: Kiểm tra hiển thị cột Trạng thái trong bảng", "N/a",
     'Hiển thị trạng thái dạng badge: "Đang làm" (màu xanh/active), "Đã đóng" (màu xám/closed), "Hoàn tất" (màu xanh/done)',
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_GUI_08", "Kiểm tra hiển thị chênh lệch đúng màu",
     "Kiểm tra hiển thị cột Chênh lệch với màu sắc phân biệt theo giá trị dương/âm",
     steps_base + "Bước 3: Kiểm tra hiển thị cột Chênh lệch trong bảng", "N/a",
     "Chênh lệch dương (+) hiển thị màu xanh, chênh lệch âm (-) hiển thị màu đỏ, bằng 0 hiển thị bình thường",
     "", "", "", "", "", "", "", "", "Bách"],
]
row = write_tests(ws, row, gui_tests)

# ============ BỘ LỌC ============
row = write_section(ws, row, "KIỂM TRA CHỨC NĂNG BỘ LỌC LỊCH SỬ HOẠT ĐỘNG", loc_fill, loc_font)

loc_tests = [
    ["LSHD_LOC_01", "Tìm kiếm theo tên nhân viên hợp lệ",
     "Kiểm tra chức năng tìm kiếm thành công theo tên nhân viên hợp lệ",
     steps_base + 'Bước 3: Nhập tên nhân viên hợp lệ vào ô tìm kiếm\nBước 4: Kiểm tra kết quả trả về',
     'Tìm kiếm: "Nguyễn Văn A"',
     "Hiển thị danh sách lịch sử hoạt động có tên nhân viên chứa từ khóa tìm kiếm",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_LOC_02", "Tìm kiếm theo mã giao ca",
     "Kiểm tra chức năng tìm kiếm thành công theo mã giao ca",
     steps_base + 'Bước 3: Nhập mã giao ca hợp lệ vào ô tìm kiếm\nBước 4: Kiểm tra kết quả trả về',
     'Tìm kiếm: "GC00001"',
     "Hiển thị bản ghi giao ca có mã GC00001",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_LOC_03", "Tìm kiếm theo tên ca làm",
     "Kiểm tra chức năng tìm kiếm thành công theo tên ca làm",
     steps_base + 'Bước 3: Nhập tên ca làm vào ô tìm kiếm\nBước 4: Kiểm tra kết quả trả về',
     'Tìm kiếm: "Sáng"',
     "Hiển thị danh sách bản ghi có tên ca chứa từ khóa \"Sáng\"",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_LOC_04", "Tìm kiếm theo mã hóa đơn (ADMIN_CASH)",
     "Kiểm tra chức năng tìm kiếm thành công theo mã hóa đơn cho bản ghi ADMIN_CASH",
     steps_base + 'Bước 3: Nhập mã hóa đơn vào ô tìm kiếm\nBước 4: Kiểm tra kết quả trả về',
     'Tìm kiếm: "HD00001"',
     "Hiển thị bản ghi ADMIN_CASH có mã hóa đơn chứa từ khóa tìm kiếm",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_LOC_05", "Tìm kiếm không có kết quả",
     "Kiểm tra chức năng tìm kiếm khi nhập từ khóa không tồn tại trong hệ thống",
     steps_base + 'Bước 3: Nhập từ khóa không tồn tại vào ô tìm kiếm\nBước 4: Kiểm tra kết quả trả về',
     'Tìm kiếm: "xyzkhongtontai123"',
     "Bảng hiển thị trống hoặc thông báo không có dữ liệu phù hợp",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_LOC_06", "Lọc theo khoảng ngày hợp lệ",
     "Kiểm tra chức năng lọc thành công theo khoảng ngày hợp lệ (Từ ngày < Đến ngày)",
     steps_base + 'Bước 3: Chọn Từ ngày và Đến ngày hợp lệ\nBước 4: Kiểm tra kết quả trả về',
     'Từ ngày: "01/03/2026"\nĐến ngày: "31/03/2026"',
     "Hiển thị danh sách lịch sử hoạt động trong khoảng thời gian đã chọn",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_LOC_07", "Lọc theo ngày cụ thể (cùng ngày)",
     "Kiểm tra chức năng lọc thành công khi Từ ngày = Đến ngày (cùng ngày)",
     steps_base + 'Bước 3: Chọn Từ ngày = Đến ngày cùng 1 ngày\nBước 4: Kiểm tra kết quả trả về',
     'Từ ngày: "15/03/2026"\nĐến ngày: "15/03/2026"',
     "Hiển thị danh sách lịch sử hoạt động chỉ trong ngày 15/03/2026",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_LOC_08", "Lọc theo khoảng ngày không có dữ liệu",
     "Kiểm tra chức năng lọc khi chọn khoảng ngày không có bản ghi nào",
     steps_base + 'Bước 3: Chọn khoảng ngày không có dữ liệu\nBước 4: Kiểm tra kết quả trả về',
     'Từ ngày: "01/01/2020"\nĐến ngày: "02/01/2020"',
     "Bảng hiển thị trống hoặc thông báo không có dữ liệu",
     "", "", "", "", "", "", "", "", "Bách"],
]
row = write_tests(ws, row, loc_tests)

# ============ PVTD - TÌM KIẾM ============
row = write_section(ws, row, "KỸ THUẬT PHÂN VÙNG TƯƠNG ĐƯƠNG - TÌM KIẾM", pvtd_fill, pvtd_font)

pvtd_search = [
    ["LSHD_PVTD_01", "Tìm kiếm từ khóa là chuỗi chữ hợp lệ (phân vùng: chữ cái)",
     "Kiểm tra tìm kiếm khi nhập từ khóa chỉ chứa chữ cái (phân vùng hợp lệ: chuỗi chữ)",
     steps_base + 'Bước 3: Nhập từ khóa chỉ chứa chữ cái vào ô tìm kiếm\nBước 4: Kiểm tra kết quả trả về',
     'Tìm kiếm: "Sáng"',
     "Hệ thống lọc và hiển thị đúng các bản ghi chứa từ khóa \"Sáng\"",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_PVTD_02", "Tìm kiếm từ khóa là chuỗi số hợp lệ (phân vùng: số)",
     "Kiểm tra tìm kiếm khi nhập từ khóa chỉ chứa số (phân vùng hợp lệ: chuỗi số)",
     steps_base + 'Bước 3: Nhập từ khóa chỉ chứa số vào ô tìm kiếm\nBước 4: Kiểm tra kết quả trả về',
     'Tìm kiếm: "00001"',
     "Hệ thống lọc và hiển thị đúng các bản ghi có mã giao ca/mã hóa đơn chứa \"00001\"",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_PVTD_03", "Tìm kiếm từ khóa chứa chữ + số (phân vùng: hỗn hợp)",
     "Kiểm tra tìm kiếm khi nhập từ khóa chứa cả chữ và số (phân vùng hợp lệ: hỗn hợp)",
     steps_base + 'Bước 3: Nhập từ khóa chứa chữ và số vào ô tìm kiếm\nBước 4: Kiểm tra kết quả trả về',
     'Tìm kiếm: "GC00001"',
     "Hệ thống lọc và hiển thị đúng bản ghi có mã giao ca GC00001",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_PVTD_04", "Tìm kiếm từ khóa chứa ký tự đặc biệt (phân vùng: ký tự đặc biệt)",
     "Kiểm tra tìm kiếm khi nhập từ khóa chứa ký tự đặc biệt (phân vùng: đặc biệt)",
     steps_base + 'Bước 3: Nhập từ khóa chứa ký tự đặc biệt vào ô tìm kiếm\nBước 4: Kiểm tra kết quả trả về',
     'Tìm kiếm: "@#$%"',
     "Bảng hiển thị trống hoặc không có kết quả phù hợp (không xảy ra lỗi hệ thống)",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_PVTD_05", "Tìm kiếm từ khóa có khoảng trắng đầu (phân vùng: trim)",
     "Kiểm tra tìm kiếm khi nhập từ khóa có khoảng trắng ở đầu (phân vùng: trim xử lý)",
     steps_base + 'Bước 3: Nhập từ khóa có khoảng trắng ở đầu vào ô tìm kiếm\nBước 4: Kiểm tra kết quả trả về',
     'Tìm kiếm: "     GC00001"',
     "Hệ thống tự động trim và lọc đúng bản ghi có mã giao ca GC00001",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_PVTD_06", "Tìm kiếm từ khóa toàn khoảng trắng (phân vùng: blank)",
     "Kiểm tra tìm kiếm khi nhập từ khóa chỉ toàn khoảng trắng (phân vùng: blank/rỗng)",
     steps_base + 'Bước 3: Nhập toàn khoảng trắng vào ô tìm kiếm\nBước 4: Kiểm tra kết quả trả về',
     'Tìm kiếm: "       "',
     "Hệ thống hiển thị toàn bộ danh sách (coi như không lọc) hoặc bỏ qua từ khóa trắng",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_PVTD_07", "Lọc ngày tuNgay < denNgay (phân vùng hợp lệ: khoảng ngày đúng)",
     "Kiểm tra lọc ngày khi Từ ngày nhỏ hơn Đến ngày (phân vùng hợp lệ: khoảng ngày thuận)",
     steps_base + 'Bước 3: Chọn Từ ngày < Đến ngày\nBước 4: Kiểm tra kết quả trả về',
     'Từ ngày: "01/03/2026"\nĐến ngày: "15/03/2026"',
     "Hiển thị danh sách bản ghi trong khoảng 01/03 đến 15/03/2026",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_PVTD_08", "Lọc ngày tuNgay > denNgay (phân vùng không hợp lệ: khoảng ngày ngược)",
     "Kiểm tra lọc ngày khi Từ ngày lớn hơn Đến ngày (phân vùng không hợp lệ: ngày ngược)",
     steps_base + 'Bước 3: Chọn Từ ngày > Đến ngày\nBước 4: Kiểm tra kết quả trả về',
     'Từ ngày: "31/03/2026"\nĐến ngày: "01/03/2026"',
     "Bảng hiển thị trống hoặc hệ thống cảnh báo khoảng ngày không hợp lệ",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_PVTD_09", "Lọc ngày chỉ nhập tuNgay, bỏ trống denNgay (phân vùng: thiếu dữ liệu)",
     "Kiểm tra lọc ngày khi chỉ nhập Từ ngày, bỏ trống Đến ngày (phân vùng: thiếu 1 trường)",
     steps_base + 'Bước 3: Chỉ chọn Từ ngày, bỏ trống Đến ngày\nBước 4: Kiểm tra kết quả trả về',
     'Từ ngày: "01/03/2026"\nĐến ngày: (trống)',
     "Hệ thống hiển thị bản ghi từ ngày 01/03/2026 đến hiện tại hoặc lọc bình thường",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_PVTD_10", "Lọc ngày chỉ nhập denNgay, bỏ trống tuNgay (phân vùng: thiếu dữ liệu)",
     "Kiểm tra lọc ngày khi chỉ nhập Đến ngày, bỏ trống Từ ngày (phân vùng: thiếu 1 trường)",
     steps_base + 'Bước 3: Chỉ chọn Đến ngày, bỏ trống Từ ngày\nBước 4: Kiểm tra kết quả trả về',
     'Từ ngày: (trống)\nĐến ngày: "31/03/2026"',
     "Hệ thống hiển thị bản ghi từ đầu đến ngày 31/03/2026 hoặc lọc bình thường",
     "", "", "", "", "", "", "", "", "Bách"],
]
row = write_tests(ws, row, pvtd_search)

# ============ GTB - TÌM KIẾM ============
row = write_section(ws, row, "KỸ THUẬT GIÁ TRỊ BIÊN - TÌM KIẾM", gtb_fill, gtb_font)

gtb_search = [
    ["LSHD_GTB_01", "Tìm kiếm từ khóa 1 ký tự (biên dưới min)",
     "Kiểm tra tìm kiếm khi nhập từ khóa chỉ 1 ký tự - giá trị biên dưới tối thiểu",
     steps_base + 'Bước 3: Nhập từ khóa 1 ký tự vào ô tìm kiếm\nBước 4: Kiểm tra kết quả trả về',
     'Tìm kiếm: "G"',
     "Hệ thống lọc và hiển thị các bản ghi có chứa ký tự \"G\" (trong mã giao ca, tên NV...)",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_GTB_02", "Tìm kiếm từ khóa 2 ký tự (cận biên dưới)",
     "Kiểm tra tìm kiếm khi nhập từ khóa 2 ký tự - giá trị cận biên dưới",
     steps_base + 'Bước 3: Nhập từ khóa 2 ký tự vào ô tìm kiếm\nBước 4: Kiểm tra kết quả trả về',
     'Tìm kiếm: "GC"',
     "Hệ thống lọc và hiển thị các bản ghi có mã giao ca bắt đầu bằng \"GC\"",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_GTB_03", "Tìm kiếm từ khóa 100 ký tự (biên trên - chuỗi dài)",
     "Kiểm tra tìm kiếm khi nhập từ khóa rất dài 100 ký tự - giá trị biên trên",
     steps_base + 'Bước 3: Nhập từ khóa 100 ký tự vào ô tìm kiếm\nBước 4: Kiểm tra kết quả trả về',
     'Tìm kiếm: "a" x 100 lần (chuỗi 100 ký tự "a")',
     "Bảng hiển thị trống (không có bản ghi phù hợp), hệ thống không xảy ra lỗi",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_GTB_04", "Tìm kiếm từ khóa chuỗi rỗng (biên 0)",
     'Kiểm tra tìm kiếm khi ô tìm kiếm rỗng - giá trị biên 0 ký tự',
     steps_base + 'Bước 3: Để trống ô tìm kiếm (xóa hết nội dung)\nBước 4: Kiểm tra kết quả trả về',
     'Tìm kiếm: ""',
     "Hệ thống hiển thị toàn bộ danh sách lịch sử hoạt động (không lọc)",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_GTB_05", "Lọc ngày cùng ngày hôm nay (biên: cùng ngày)",
     "Kiểm tra lọc ngày khi Từ ngày = Đến ngày = ngày hôm nay - giá trị biên cùng ngày",
     steps_base + 'Bước 3: Chọn Từ ngày = Đến ngày = ngày hôm nay\nBước 4: Kiểm tra kết quả trả về',
     'Từ ngày: ngày hôm nay\nĐến ngày: ngày hôm nay',
     "Hiển thị danh sách bản ghi chỉ trong ngày hôm nay",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_GTB_06", "Lọc ngày từ ngày đầu tiên có dữ liệu (biên dưới ngày)",
     "Kiểm tra lọc ngày từ ngày đầu tiên có bản ghi trong hệ thống - giá trị biên dưới",
     steps_base + 'Bước 3: Chọn Từ ngày = ngày đầu tiên có dữ liệu\nBước 4: Kiểm tra kết quả trả về',
     'Từ ngày: ngày đầu tiên có dữ liệu (VD: "01/01/2026")\nĐến ngày: "31/03/2026"',
     "Hiển thị toàn bộ bản ghi từ ngày đầu tiên đến ngày chọn",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_GTB_07", "Lọc ngày đến ngày cuối cùng có dữ liệu (biên trên ngày)",
     "Kiểm tra lọc ngày đến ngày cuối cùng có bản ghi trong hệ thống - giá trị biên trên",
     steps_base + 'Bước 3: Chọn Đến ngày = ngày cuối cùng có dữ liệu\nBước 4: Kiểm tra kết quả trả về',
     'Từ ngày: "01/01/2026"\nĐến ngày: ngày cuối cùng có dữ liệu',
     "Hiển thị toàn bộ bản ghi đến ngày cuối cùng",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_GTB_08", "Lọc ngày khoảng 1 ngày liền kề (biên khoảng ngày nhỏ nhất)",
     "Kiểm tra lọc ngày khi Từ ngày và Đến ngày liền kề nhau (khoảng 1 ngày) - biên khoảng ngày min",
     steps_base + 'Bước 3: Chọn Từ ngày liền kề Đến ngày\nBước 4: Kiểm tra kết quả trả về',
     'Từ ngày: "14/03/2026"\nĐến ngày: "15/03/2026"',
     "Hiển thị bản ghi trong khoảng 2 ngày 14-15/03/2026",
     "", "", "", "", "", "", "", "", "Bách"],
]
row = write_tests(ws, row, gtb_search)

# ============ PVTD - HIỂN THỊ DỮ LIỆU ============
row = write_section(ws, row, "KỸ THUẬT PHÂN VÙNG TƯƠNG ĐƯƠNG - HIỂN THỊ DỮ LIỆU", pvtd_dl_fill, pvtd_dl_font)

pvtd_dl = [
    ["LSHD_PVTD_DL_01", "Hiển thị bản ghi loại GIAO_CA (phân vùng: giao ca nhân viên)",
     "Kiểm tra hiển thị đúng bản ghi loại GIAO_CA - giao ca của nhân viên (phân vùng: loại bản ghi GIAO_CA)",
     steps_base + 'Bước 3: Quan sát bảng dữ liệu, xác định bản ghi loại GIAO_CA\nBước 4: Kiểm tra hiển thị đầy đủ thông tin',
     "Dữ liệu có sẵn: bản ghi giao ca của nhân viên (loaiBanGhi = GIAO_CA)",
     "Hiển thị đúng: tên nhân viên, tên ca làm, thời gian mở/đóng, tiền mặt, chênh lệch, trạng thái (Đang làm/Đã đóng)",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_PVTD_DL_02", "Hiển thị bản ghi loại ADMIN_CASH (phân vùng: admin thu tiền)",
     "Kiểm tra hiển thị đúng bản ghi loại ADMIN_CASH - admin thu tiền mặt (phân vùng: loại bản ghi ADMIN_CASH)",
     steps_base + 'Bước 3: Quan sát bảng dữ liệu, xác định bản ghi loại ADMIN_CASH\nBước 4: Kiểm tra hiển thị đầy đủ thông tin',
     "Dữ liệu có sẵn: bản ghi admin thu tiền mặt (loaiBanGhi = ADMIN_CASH)",
     'Hiển thị đúng: tên tài khoản + mã hóa đơn, ca = "ADMIN" badge, tiền mặt, trạng thái = "Hoàn tất"',
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_PVTD_DL_03", "Hiển thị trạng thái Đang làm (phân vùng: trangThai=0)",
     'Kiểm tra hiển thị đúng trạng thái "Đang làm" cho ca đang mở (phân vùng: trạng thái đang hoạt động)',
     steps_base + 'Bước 3: Quan sát bản ghi có ca đang mở (chưa kết thúc)\nBước 4: Kiểm tra cột Trạng thái',
     "Dữ liệu có sẵn: ca làm đang mở (trangThai = 0, thoiGianKetCa = null)",
     'Hiển thị badge "Đang làm" với màu active, cột Đóng hiển thị trống',
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_PVTD_DL_04", "Hiển thị trạng thái Đã đóng (phân vùng: trangThai=1)",
     'Kiểm tra hiển thị đúng trạng thái "Đã đóng" cho ca đã kết thúc (phân vùng: trạng thái đã đóng)',
     steps_base + 'Bước 3: Quan sát bản ghi có ca đã kết thúc\nBước 4: Kiểm tra cột Trạng thái',
     "Dữ liệu có sẵn: ca làm đã đóng (trangThai = 1, thoiGianKetCa có giá trị)",
     'Hiển thị badge "Đã đóng" với màu closed, cột Đóng hiển thị thời gian kết thúc ca',
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_PVTD_DL_05", "Hiển thị chênh lệch dương (phân vùng: chênh lệch > 0)",
     "Kiểm tra hiển thị đúng chênh lệch dương (+) khi tiền dư (phân vùng: chênh lệch dương)",
     steps_base + 'Bước 3: Quan sát bản ghi có chênh lệch dương\nBước 4: Kiểm tra cột Chênh lệch',
     "Dữ liệu có sẵn: bản ghi có tienDauCaNhap > tienBanGiaoDuKien (tiền dư)",
     "Hiển thị giá trị chênh lệch dương (+) với màu xanh lá",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_PVTD_DL_06", "Hiển thị chênh lệch âm (phân vùng: chênh lệch < 0)",
     "Kiểm tra hiển thị đúng chênh lệch âm (-) khi tiền thiếu (phân vùng: chênh lệch âm)",
     steps_base + 'Bước 3: Quan sát bản ghi có chênh lệch âm\nBước 4: Kiểm tra cột Chênh lệch',
     "Dữ liệu có sẵn: bản ghi có tienDauCaNhap < tienBanGiaoDuKien (tiền thiếu)",
     "Hiển thị giá trị chênh lệch âm (-) với màu đỏ",
     "", "", "", "", "", "", "", "", "Bách"],
]
row = write_tests(ws, row, pvtd_dl)

# ============ GTB - HIỂN THỊ DỮ LIỆU ============
row = write_section(ws, row, "KỸ THUẬT GIÁ TRỊ BIÊN - HIỂN THỊ DỮ LIỆU", gtb_dl_fill, gtb_dl_font)

gtb_dl = [
    ["LSHD_GTB_DL_01", "Hiển thị tiền mặt = 0 (biên dưới tiền)",
     "Kiểm tra hiển thị đúng khi giá trị tiền mặt trong ca = 0 - giá trị biên dưới",
     steps_base + 'Bước 3: Quan sát bản ghi có tiền mặt = 0\nBước 4: Kiểm tra cột Tiền mặt',
     "Dữ liệu có sẵn: bản ghi có tienMatTrongCa = 0",
     'Hiển thị "0" hoặc "0 đ" đúng định dạng trong cột Tiền mặt',
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_GTB_DL_02", "Hiển thị tiền mặt số lớn (biên trên tiền)",
     "Kiểm tra hiển thị đúng khi giá trị tiền mặt rất lớn (VD: 999,999,999) - giá trị biên trên",
     steps_base + 'Bước 3: Quan sát bản ghi có tiền mặt số lớn\nBước 4: Kiểm tra cột Tiền mặt',
     "Dữ liệu có sẵn: bản ghi có tienMatTrongCa = 999,999,999",
     "Hiển thị đúng định dạng số tiền lớn với dấu phân cách hàng nghìn (999.999.999 đ), không bị tràn cột",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_GTB_DL_03", "Hiển thị chênh lệch = 0 (biên: không chênh lệch)",
     "Kiểm tra hiển thị đúng khi chênh lệch = 0 (tiền khớp hoàn toàn) - giá trị biên 0",
     steps_base + 'Bước 3: Quan sát bản ghi có chênh lệch = 0\nBước 4: Kiểm tra cột Chênh lệch',
     "Dữ liệu có sẵn: bản ghi có tienDauCaNhap = tienBanGiaoDuKien",
     'Hiển thị "0" hoặc "0 đ" với màu bình thường (không xanh, không đỏ)',
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_GTB_DL_04", "Hiển thị chênh lệch = 1 (biên dưới dương)",
     "Kiểm tra hiển thị đúng khi chênh lệch = +1 - giá trị biên dưới dương nhỏ nhất",
     steps_base + 'Bước 3: Quan sát bản ghi có chênh lệch = +1\nBước 4: Kiểm tra cột Chênh lệch',
     "Dữ liệu có sẵn: bản ghi có chênh lệch tienDauCaNhap - tienBanGiaoDuKien = 1",
     "Hiển thị +1 với màu xanh lá (dương)",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_GTB_DL_05", "Hiển thị chênh lệch = -1 (biên dưới âm)",
     "Kiểm tra hiển thị đúng khi chênh lệch = -1 - giá trị biên dưới âm nhỏ nhất",
     steps_base + 'Bước 3: Quan sát bản ghi có chênh lệch = -1\nBước 4: Kiểm tra cột Chênh lệch',
     "Dữ liệu có sẵn: bản ghi có chênh lệch tienDauCaNhap - tienBanGiaoDuKien = -1",
     "Hiển thị -1 với màu đỏ (âm)",
     "", "", "", "", "", "", "", "", "Bách"],
    ["LSHD_GTB_DL_06", "Hiển thị bảng khi không có dữ liệu (biên: 0 bản ghi)",
     "Kiểm tra hiển thị bảng khi hệ thống không có bất kỳ bản ghi lịch sử hoạt động nào - biên 0 bản ghi",
     steps_base + 'Bước 3: Truy cập màn hình khi không có dữ liệu lịch sử\nBước 4: Kiểm tra hiển thị bảng',
     "Hệ thống không có bản ghi lịch sử hoạt động nào (DB trống)",
     "Bảng hiển thị trống với thông báo \"Không có dữ liệu\" hoặc hiển thị bảng rỗng không lỗi",
     "", "", "", "", "", "", "", "", "Bách"],
]
row = write_tests(ws, row, gtb_dl)

# Row heights
for r in range(2, row):
    ws.row_dimensions[r].height = 90
ws.row_dimensions[1].height = 40
ws.freeze_panes = 'A2'

output_path = r"c:\Users\ducdu\Downloads\demo_client\TestCase_LichSuHoatDong_SevenStrike.xlsx"
wb.save(output_path)
print(f"File saved: {output_path}")
